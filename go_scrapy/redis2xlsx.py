import re
import time

import redis
import requests
import os
import threading
from openpyxl.drawing import image
from openpyxl import Workbook

r = redis.Redis(host='localhost', port=6379, decode_responses=True)

def mkdir(path):
    isExists = os.path.exists(path)
    if not isExists:
        os.makedirs(path)
        print('Directory create successfully!')
        return True
    else:
        print('Directory already exists!')
        return False

def filter(qipu_dict, ws2, Index):
    name = qipu_dict['name']
    qtypename = qipu_dict['qtypename']
    isblackfirst = qipu_dict['blackfirst']
    prepos = qipu_dict['prepos']
    levelname = qipu_dict['levelname']
    answers = qipu_dict['answers']
    vv = qipu_dict['vv']

    ws2.cell(Index, 3).value = name
    ws2.cell(Index, 4).value = qtypename
    ws2.cell(Index, 5).value = levelname
    prepos2 = []
    prepos_x = []
    prepos_y = []
    if (vv % 3 != 0):
        for x in prepos[0]:
            prepos_x.append(x[1] + x[0])
        prepos2.append(prepos_x)
        for y in prepos[1]:
            prepos_y.append(y[1] + y[0])
        prepos2.append(prepos_y)
    else:
        prepos2 = prepos
    ws2.cell(Index, 7).value = str(prepos2)[1:-1].replace("[","{").replace("]","}").replace("\'","\"")
    answer_type1 = []
    # answer_type2 = []
    answer_type3 = []
    for answer in answers:
        if answer["ty"] == 1 and answer["st"] == 2:
            answer_type1.append(filter_answer(answer["pts"]))
        # elif answer["ty"] == 2:
        #     answer_type2.append(filter_answer(answer["pts"]))
        elif answer["ty"] == 3 and answer["st"] == 2:
            answer_type3.append(filter_answer(answer["pts"]))
        else:
            continue
    ws2.cell(Index, 8).value = str(answer_type1)[1:-1].replace("[","{").replace("]","}").replace("\'","\"")
    # ws2.cell(Index, 9).value = str(answer_type2)[1:-1].replace("[","{").replace("]","}").replace("\'","\"")
    ws2.cell(Index, 9).value =str(answer_type3)[1:-1].replace("[","{").replace("]","}").replace("\'","\"")
    if isblackfirst:
        ws2.cell(Index, 6).value = "0:黑色"
    else:
        ws2.cell(Index, 6).value = "1:白色"


def filter_answer(pts):
    answerlist = []
    for x in pts:
        answerlist.append(x["p"])
    return answerlist



def get_qipu(Step):
    I = 2
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.cell(1, 1).value = "棋谱图"
    ws1.cell(1, 2).value = "棋谱id"
    ws1.cell(1, 3).value = "棋谱名"
    ws1.cell(1, 4).value = "棋谱类型"
    ws1.cell(1, 5).value = "棋谱等级"
    ws1.cell(1, 6).value = "黑白先手"
    ws1.cell(1, 7).value = "棋谱"
    ws1.cell(1, 8).value = "正确答案"
    ws1.cell(1, 9).value = "错位答案"
    qipuids = r.keys("/{}K*".format(Step))
    for i in qipuids:
        try:
            if r.sismember("un_success_content", i) == 0:
                image_path = r.hget(i, 'img')
                r2 = re.compile('\d{3,}')
                qipu_id = r2.search(i).group()
                print(qipu_id)
                qipu_content = r.hget(i, 'content')
                add_excel(image_path, qipu_id, qipu_content, I, ws1)
                I += 1
            else:
                print("jump,{}".format(i))
                continue
        except Exception:
            print("Exception jump,{}".format(i))
            continue
    wb1.save('{}K.xlsx'.format(Step))


def add_excel(image_path, qipu_id, qipu_content, Index, ws1):
    image_bytes = requests.get(image_path).content
    data_stream = image.BytesIO(image_bytes)
    im = image.Image(data_stream)
    c = ws1.column_dimensions['A']
    # 设置行高列宽公式里面的96为图片的水平和垂直分辨率。即所谓的dpi。
    c.width = im.width * 12 / 96
    r = ws1.row_dimensions[Index]
    r.height = im.height * 72 / 96
    CellStr = "A" + str(Index)
    # 第一列 图片
    ws1.add_image(im, CellStr)
    # 第二列 棋谱id
    ws1.cell(Index, 2).value = qipu_id
    r = re.compile("(?<=(var g_qq = )).*(?=(;var taskinfo))")
    String = r.search(qipu_content)
    if String != None:
        String2 = String.group().replace("false", "False")
        String3 = String2.replace("true", "True")
        String4 = String3.replace("null", "None")
        qipu_dict = eval(String4)
        filter(qipu_dict, ws1, Index)




if __name__ == '__main__':
    for index in range(13,14):
        get_qipu(index)
    # L1 = [1,2]
    # L2 = [2,3]
    # print(set(r.keys("/{}K*".format(6))))
    # print(list(set(L1).difference(set(L2))))
    r.close()
    # thread = threading.Thread(target= a, args=(10, ))
    # thread.start()
    # thread.join()
    # print("start")
